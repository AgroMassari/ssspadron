import os
import time
import openpyxl
from sss_beneficiarios_hospitales.data import DataBeneficiariosSSSHospital

# ============================================================
# CONFIGURACIÓN
# ============================================================

ARCHIVO_EXCEL = "estadistico_emergencias_v9_copia.xlsx"

# Primera fila de pacientes
FILA_INICIO = 11

# Columnas
COLUMNA_DNI = 1          # A
COLUMNA_OBRA_SOCIAL = 7 # G

# Guardar cada X consultas
GUARDAR_CADA = 20

# Espera entre consultas
ESPERA = 2

usuario = os.environ["SSS_USER"]
password = os.environ["SSS_PASSWORD"]

# ============================================================
# CONEXIÓN SSSALUD
# ============================================================

sss = DataBeneficiariosSSSHospital(
    user=usuario,
    password=password
)

# No guardar respuestas individuales
sss._save_response = lambda filename, resp: None

# ============================================================
# ABRIR EXCEL
# ============================================================

wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
ws = wb.active

ultima_fila = ws.max_row

print("")
print("====================================================")
print("       PROCESAMIENTO ESTADÍSTICO SSSALUD")
print("====================================================")
print("")
print("Archivo:", ARCHIVO_EXCEL)
print("Primera fila:", FILA_INICIO)
print("Última fila:", ultima_fila)
print("Columna DNI: A")
print("Columna Obra Social: G")
print("")
print("IMPORTANTE:")
print("- Las celdas G con información NO se consultan.")
print("- Se guardará el Excel cada", GUARDAR_CADA, "consultas.")
print("- Espera entre consultas:", ESPERA, "segundos.")
print("")
print("====================================================")
print("")

consultas = 0
procesadas = 0
saltadas = 0
no_afiliados = 0
afiliados = 0
errores = 0
sin_dni = 0

try:

    for fila in range(FILA_INICIO, ultima_fila + 1):

        dni = ws.cell(fila, COLUMNA_DNI).value
        valor_actual = ws.cell(fila, COLUMNA_OBRA_SOCIAL).value

        # ====================================================
        # DNI VACÍO
        # ====================================================

        if dni is None or str(dni).strip() == "":

            sin_dni += 1

            if valor_actual is None or str(valor_actual).strip() == "":
                ws.cell(fila, COLUMNA_OBRA_SOCIAL).value = "SIN DNI"

            continue

        dni = str(dni).strip()

        # ====================================================
        # SI G YA TIENE INFORMACIÓN, NO CONSULTAR
        # ====================================================

        if valor_actual is not None and str(valor_actual).strip() != "":

            saltadas += 1
            continue

        # ====================================================
        # MOSTRAR PROGRESO
        # ====================================================

        procesadas += 1

        print("")
        print("----------------------------------------------------")
        print(f"Fila {fila} | DNI {dni}")
        print(f"Progreso: {fila - FILA_INICIO + 1}/{ultima_fila - FILA_INICIO + 1}")
        print("Consultando SSSalud...")
        print("----------------------------------------------------")

        # ====================================================
        # CONSULTAR
        # ====================================================

        try:

            resultado = sss.query(dni)

            ok = resultado.get("ok")
            datos = resultado.get("resultados", {})

            afiliado = datos.get("afiliado")
            tablas = datos.get("tablas", [])

            print("OK:", ok)
            print("Afiliado:", afiliado)

            # =================================================
            # NO AFILIADO
            # =================================================

            if afiliado is False:

                resultado_final = "NO AFILIADO"

                no_afiliados += 1

                print("Resultado: NO AFILIADO")

            # =================================================
            # AFILIADO
            # =================================================

            elif afiliado is True:

                codigo_obra_social = None
                denominacion_obra_social = None

                # Buscar tabla AFILIADO
                for tabla in tablas:

                    nombre_tabla = str(tabla.get("name", "")).strip().upper()

                    if nombre_tabla != "AFILIADO":
                        continue

                    data = tabla.get("data", {})

                    codigo_obra_social = data.get(
                        "Código de Obra Social"
                    )

                    denominacion_obra_social = data.get(
                        "Denominación Obra Social"
                    )

                    break

                # ---------------------------------------------
                # FORMAR RESULTADO
                # ---------------------------------------------

                if codigo_obra_social and denominacion_obra_social:

                    resultado_final = (
                        f"{codigo_obra_social} - "
                        f"{denominacion_obra_social}"
                    )

                elif denominacion_obra_social:

                    resultado_final = str(denominacion_obra_social)

                elif codigo_obra_social:

                    resultado_final = str(codigo_obra_social)

                else:

                    resultado_final = "AFILIADO - SIN OBRA SOCIAL IDENTIFICADA"

                afiliados += 1

                print("Resultado:", resultado_final)

            # =================================================
            # RESPUESTA INESPERADA
            # =================================================

            else:

                resultado_final = "ERROR CONSULTA"

                errores += 1

                print("Resultado: ERROR CONSULTA")

            # =================================================
            # ESCRIBIR EN G
            # =================================================

            ws.cell(
                fila,
                COLUMNA_OBRA_SOCIAL
            ).value = resultado_final

            consultas += 1

        except Exception as e:

            print("")
            print("ERROR DE CONSULTA:")
            print(str(e))

            ws.cell(
                fila,
                COLUMNA_OBRA_SOCIAL
            ).value = "ERROR CONSULTA"

            errores += 1
            consultas += 1

        # ====================================================
        # GUARDADO PERIÓDICO
        # ====================================================

        if consultas > 0 and consultas % GUARDAR_CADA == 0:

            print("")
            print(">>> GUARDANDO PROGRESO...")
            wb.save(ARCHIVO_EXCEL)
            print(">>> PROGRESO GUARDADO")
            print("")

        # ====================================================
        # ESPERA
        # ====================================================

        time.sleep(ESPERA)

except KeyboardInterrupt:

    print("")
    print("")
    print("====================================================")
    print("          PROCESO INTERRUMPIDO")
    print("====================================================")
    print("")
    print("Guardando progreso antes de salir...")
    wb.save(ARCHIVO_EXCEL)
    print("Progreso guardado correctamente.")
    print("")
    print("Podés volver a ejecutar el script para continuar.")
    print("")

# ============================================================
# GUARDADO FINAL
# ============================================================

wb.save(ARCHIVO_EXCEL)

print("")
print("====================================================")
print("             PROCESO TERMINADO")
print("====================================================")
print("")
print("Archivo:", ARCHIVO_EXCEL)
print("")
print("Resumen:")
print("Consultas realizadas:", consultas)
print("Pacientes procesados:", procesadas)
print("Filas saltadas:", saltadas)
print("Afiliados:", afiliados)
print("No afiliados:", no_afiliados)
print("Errores:", errores)
print("Sin DNI:", sin_dni)
print("")
print("Excel guardado correctamente.")
print("====================================================")