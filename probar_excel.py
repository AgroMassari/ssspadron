import os
from openpyxl import load_workbook
from sss_beneficiarios_hospitales.data import DataBeneficiariosSSSHospital

ARCHIVO = "estadistico_emergencias_v9_copia.xlsx"
HOJA = "Sheet0"
FILA = 12

usuario = os.environ["SSS_USER"]
password = os.environ["SSS_PASSWORD"]

wb = load_workbook(ARCHIVO)
ws = wb[HOJA]

dni = ws.cell(FILA, 1).value

print(f"DNI de la fila {FILA}: {dni}")
print("Consultando SSSalud...")

sss = DataBeneficiariosSSSHospital(
    user=usuario,
    password=password
)

# Evita el problema de codificación de la librería
sss._save_response = lambda filename, resp: None

try:
    resultado = sss.query(str(dni))

    if not resultado.get("ok"):
        print("ERROR: La consulta no fue exitosa.")
        print(resultado)
        wb.close()
        raise SystemExit(1)

    print("Consulta correcta.")

    datos = resultado.get("resultados", {})
    tablas = datos.get("tablas", [])

    obra_social = None

    for tabla in tablas:
        if tabla.get("name") == "AFILIADO":
            data = tabla.get("data", {})
            obra_social = data.get("Denominación Obra Social")
            codigo = data.get("Código de Obra Social")

            print(f"Código obra social: {codigo}")
            print(f"Obra social: {obra_social}")
            break

    if obra_social:
        ws.cell(FILA, 7).value = obra_social
        wb.save(ARCHIVO)

        print("")
        print(f"EXITO: G{FILA} fue actualizada.")
        print(f"Valor escrito: {obra_social}")
    else:
        print("")
        print("No se encontró una obra social en la respuesta.")
        print("No se modificó el Excel.")

except Exception as e:
    print("")
    print("ERROR:")
    print(type(e).__name__)
    print(str(e))

finally:
    wb.close()
