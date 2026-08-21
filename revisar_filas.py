from openpyxl import load_workbook

archivo = "estadistico_emergencias_v9_copia.xlsx"

wb = load_workbook(archivo, read_only=True, data_only=False)
ws = wb["Sheet0"]

print("Filas 10 a 30:")
print("")

for fila in range(10, 31):
    valores = [ws.cell(fila, col).value for col in range(1, 10)]
    print(f"Fila {fila}: {valores}")

wb.close()
