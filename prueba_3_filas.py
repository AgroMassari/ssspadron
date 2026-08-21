import os
import time
import openpyxl
from sss_beneficiarios_hospitales.data import DataBeneficiariosSSSHospital

# ============================================================
# CONFIGURACIÓN
# ============================================================

ARCHIVO_EXCEL = "estadistico_emergencias_v9_copia.xlsx"

# Filas que vamos a probar
FILAS_PRUEBA = [12, 13, 14]

# Excel:
# A = 1 -> DNI
# G = 7 -> Obra Social
COLUMNA_DNI = 1
COLUMNA_OBRA_SOCIAL = 7

usuario = os.environ["SSS_USER"]
password = os.environ["SSS_PASSWORD"]

# ============================================================
# CONEXIÓN SSSALUD
# ============================================================

sss = DataBeneficiariosSSSHospital(
    user=usuario,
    password=password
)

# No guardar respuestas individuales
sss._save_response = lambda filename, resp: None

# ============================================================
# ABRIR EXCEL
# ============================================================

wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
ws = wb.active

print("")
print("==============================================")
print("     PRUEBA SSSALUD - 3 PACIENTES")
print("==============================================")
print("")
print("Archivo:", ARCHIVO_EXCEL)
print("Filas:", FILAS_PRUEBA)
print("")

# ============================================================
# PROCESAR
# ============================================================

for fila in FILAS_PRUEBA:

    dni = ws.cell(fila, COLUMNA_DNI).value
    valor_actual = ws.cell(fila, COLUMNA_OBRA_SOCIAL).value

    print("----------------------------------------------")
    print("Fila:", fila)
    print("DNI:", dni)
    print("G actual:", valor_actual)

    # --------------------------------------------------------
    # VALIDAR DNI
    # --------------------------------------------------------

    if dni is None or str(dni).strip() == "":
        print("Resultado: SIN DNI")
        ws.cell(fila, COLUMNA_OBRA_SOCIAL).value = "SIN DNI"
        continue

    dni = str(dni).strip()

    # --------------------------------------------------------
    # EN ESTA PRUEBA PERMITIMOS REPROCESAR ERROR CONSULTA
    # --------------------------------------------------------

    if (
        valor_actual is not None
        and str(valor_actual).strip() != ""
        and str(valor_actual).strip().upper() != "ERROR CONSULTA"
    ):
        print("Resultado: NO SE CONSULTA")
        print("Motivo: G ya tiene información válida")
        continue

    # --------------------------------------------------------
    # CONSULTA
    # --------------------------------------------------------

    try:

        print("Consultando SSSalud para DNI:", dni)

        resultado = sss.query(dni)

        print("")
        print("RESPUESTA COMPLETA:")
        print(resultado)
        print("")

        datos = resultado.get("resultados", {})

        print("OK:", resultado.get("ok"))
        print("Título:", datos.get("title"))
        print("Afiliado:", datos.get("afiliado"))
        print("")

        tablas = datos.get("tablas", [])

        print("TABLAS ENCONTRADAS:", len(tablas))
        print("")

        resultado_final = None

        # ====================================================
        # ANALIZAR TABLAS
        # ====================================================

        for tabla in tablas:

            nombre_tabla = tabla.get("name")
            data = tabla.get("data", {})

            print("==============================================")
            print("TABLA:", nombre_tabla)
            print("==============================================")

            print("Campos encontrados:")

            for clave, valor in data.items():
                print("  ", clave, "=", valor)

            print("")

            # ------------------------------------------------
            # NO AFILIADO
            # ------------------------------------------------

            if str(nombre_tabla).strip().upper() == "NO_AFILIADO":

                resultado_final = "NO AFILIADO"

                print("Detectado: NO AFILIADO")

                break

        # ====================================================
        # SI NO ES NO_AFILIADO, BUSCAMOS POSIBLE OBRA SOCIAL
        # ====================================================

        if resultado_final is None:

            for tabla in tablas:

                nombre_tabla = tabla.get("name")
                data = tabla.get("data", {})

                for clave, valor in data.items():

                    if valor is None:
                        continue

                    texto = str(valor).strip()

                    if texto == "":
                        continue

                    clave_lower = str(clave).lower()

                    if (
                        "obra social" in clave_lower
                        or "obra_social" in clave_lower
                        or "entidad" in clave_lower
                        or "agente" in clave_lower
                        or "denominacion" in clave_lower
                        or "razon social" in clave_lower
                        or "razón social" in clave_lower
                    ):
                        resultado_final = texto
                        break

                if resultado_final:
                    break

        # ====================================================
        # SI NO ENCONTRAMOS NOMBRE
        # ====================================================

        if resultado_final is None:

            if resultado.get("ok") is True:

                resultado_final = "AFILIADO - REVISAR RESPUESTA"

            else:

                resultado_final = "ERROR CONSULTA"

        # ====================================================
        # GUARDAR RESULTADO
        # ====================================================

        ws.cell(
            fila,
            COLUMNA_OBRA_SOCIAL
        ).value = resultado_final

        print("")
        print("**********************************************")
        print("RESULTADO FINAL:", resultado_final)
        print("**********************************************")
        print("")

    except Exception as e:

        print("")
        print("ERROR:")
        print(str(e))
        print("")

        ws.cell(
            fila,
            COLUMNA_OBRA_SOCIAL
        ).value = "ERROR CONSULTA"

    # Esperar entre consultas
    time.sleep(2)

# ============================================================
# GUARDAR EXCEL
# ============================================================

wb.save(ARCHIVO_EXCEL)

print("")
print("==============================================")
print("          PRUEBA TERMINADA")
print("==============================================")
print("")
print("Excel guardado:")
print(ARCHIVO_EXCEL)
print("")
print("Revisamos las filas 12, 13 y 14.")
print("")