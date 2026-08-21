from openpyxl import load_workbook

archivo = "estadistico_emergencias_v9_copia.xlsx"

wb = load_workbook(archivo, read_only=True, data_only=False)

print("HOJAS DEL ARCHIVO:")
for ws in wb.worksheets:
    print(f"- {ws.title}")

    print("DIMENSION:", ws.max_row, "filas x", ws.max_column, "columnas")

    print("PRIMERAS 10 FILAS:")
    for fila in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True):
        print(fila)

    print("")
    
wb.close()
