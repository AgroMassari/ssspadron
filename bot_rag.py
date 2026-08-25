"""
bot_rag.py
==========
Motor de inteligencia artificial (RAG) para el bot de WhatsApp y analizador de fojas quirúrgicas.

Formatos soportados para análisis:
  - Imagen directa: JPG, PNG, WEBP → GPT-4o Vision
  - PDF con texto extraíble         → GPT-4o + RAG (nomencladores)
  - PDF escaneado (sin texto)       → convierte a imagen → GPT-4o Vision + RAG

DEPENDENCIAS: pip install langchain langchain-openai langchain-community chromadb PyMuPDF tiktoken openai
"""

import os
import logging
from pathlib import Path

# ============================================================
# CONFIGURACIÓN
# ============================================================

BASE_DIR        = Path(__file__).resolve().parent
DOCS_DIR        = BASE_DIR / "documentos"
CHROMA_DIR      = BASE_DIR / "chroma_db"

DOCS_DIR.mkdir(exist_ok=True)
CHROMA_DIR.mkdir(exist_ok=True)

CHUNK_SIZE      = 1000
CHUNK_OVERLAP   = 150
MODEL_NAME      = "gpt-4o-mini"        # Para respuestas de texto (más económico)
MODEL_VISION    = "gpt-4o-mini"             # Para análisis de imágenes (visión completa y más económica)
EMBEDDING_MODEL = "text-embedding-3-small"

EXTENSIONES_IMAGEN = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
MIME_TIPOS = {
    ".jpg":  "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png":  "image/png",
    ".webp": "image/webp",
    ".gif":  "image/gif",
}

logger = logging.getLogger(__name__)


# ============================================================
# CARGA DE DOCUMENTOS (para indexar en ChromaDB)
# ============================================================

def _cargar_pdf(ruta: Path) -> list:
    from langchain_community.document_loaders import PyMuPDFLoader
    return PyMuPDFLoader(str(ruta)).load()


def _cargar_txt(ruta: Path) -> list:
    from langchain_community.document_loaders import TextLoader
    return TextLoader(str(ruta), encoding="utf-8").load()


def _cargar_docx(ruta: Path) -> list:
    from langchain_community.document_loaders import UnstructuredWordDocumentLoader
    return UnstructuredWordDocumentLoader(str(ruta)).load()


def _cargar_excel(ruta: Path) -> list:
    import openpyxl
    from langchain.schema import Document
    documentos = []
    # read_only=True es CLAVE para no saturar la memoria RAM
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    try:
        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            filas_texto = []
            for fila in ws.iter_rows(values_only=True):
                partes = [str(c) for c in fila if c is not None]
                if partes:
                    filas_texto.append(" | ".join(partes))
            if filas_texto:
                contenido = "=== Hoja: {} ===\n{}".format(nombre_hoja, "\n".join(filas_texto))
                documentos.append(Document(
                    page_content=contenido,
                    metadata={"source": str(ruta), "sheet": nombre_hoja},
                ))
    finally:
        wb.close()
    return documentos


def cargar_documento(ruta: Path) -> list:
    ext = ruta.suffix.lower()
    if ext == ".pdf":
        return _cargar_pdf(ruta)
    elif ext == ".txt":
        return _cargar_txt(ruta)
    elif ext in (".docx", ".doc"):
        return _cargar_docx(ruta)
    elif ext in (".xlsx", ".xls"):
        return _cargar_excel(ruta)
    else:
        logger.warning("Formato no soportado para indexar: %s", ruta.name)
        return []


# ============================================================
# BASE VECTORIAL (ChromaDB)
# ============================================================

def _obtener_embeddings():
    from langchain_openai import OpenAIEmbeddings
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("Falta la variable de entorno OPENAI_API_KEY.")
    return OpenAIEmbeddings(model=EMBEDDING_MODEL, openai_api_key=api_key)


def _obtener_vectorstore():
    from langchain_community.vectorstores import Chroma
    return Chroma(
        persist_directory=str(CHROMA_DIR),
        embedding_function=_obtener_embeddings(),
        collection_name="ssspadron_docs",
    )


def agregar_documento_a_base(ruta: Path) -> int:
    from langchain.text_splitter import RecursiveCharacterTextSplitter
    documentos = cargar_documento(ruta)
    if not documentos:
        return 0
    splitter   = RecursiveCharacterTextSplitter(chunk_size=CHUNK_SIZE, chunk_overlap=CHUNK_OVERLAP)
    fragmentos = splitter.split_documents(documentos)
    _obtener_vectorstore().add_documents(fragmentos)
    logger.info("'%s' indexado: %d fragmentos.", ruta.name, len(fragmentos))
    return len(fragmentos)


def indexar_todos_los_documentos() -> int:
    archivos_existentes = set()
    try:
        vs = _obtener_vectorstore()
        if vs._collection.count() > 0:
            # Recuperar metadatos para ver qué archivos ya están indexados
            resultados = vs._collection.get(include=["metadatas"])
            if resultados and resultados.get("metadatas"):
                for meta in resultados["metadatas"]:
                    if meta and "source" in meta:
                        # Extraer solo el nombre del archivo de la ruta fuente
                        archivos_existentes.add(Path(meta["source"]).name)
            logger.info(f"Archivos ya indexados en DB: {archivos_existentes}")
    except Exception as e:
        logger.warning(f"No se pudo verificar la colección: {e}")
        
    extensiones = {".pdf", ".txt", ".docx", ".doc", ".xlsx", ".xls"}
    total = 0
    for archivo in DOCS_DIR.iterdir():
        if archivo.suffix.lower() in extensiones:
            if archivo.name in archivos_existentes:
                logger.info("Omitiendo '%s', ya está indexado.", archivo.name)
                continue
            logger.info("Indexando nuevo archivo: '%s'", archivo.name)
            total += agregar_documento_a_base(archivo)
    return total


# ============================================================
# CADENA RAG (bot de WhatsApp)
# ============================================================

def _obtener_llm():
    from langchain_openai import ChatOpenAI
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("Falta la variable de entorno OPENAI_API_KEY.")
    return ChatOpenAI(model=MODEL_NAME, temperature=0.2, openai_api_key=api_key)


def construir_cadena_qa():
    from langchain.chains import RetrievalQA
    from langchain.prompts import PromptTemplate

    retriever = _obtener_vectorstore().as_retriever(
        search_type="similarity", search_kwargs={"k": 5}
    )

    PROMPT_SISTEMA = (
        "Sos un asistente especializado en salud, facturación médica y nomencladores. "
        "Respondé SIEMPRE en español, de forma clara y precisa. "
        "Si la información no está en los documentos disponibles, decilo claramente. "
        "No inventes datos ni códigos que no estén en el contexto."
    )
    template = (
        "{ps}\n\nContexto de los documentos:\n{{context}}\n\nPregunta: {{question}}\nRespuesta:"
    ).format(ps=PROMPT_SISTEMA)

    prompt = PromptTemplate(input_variables=["context", "question"], template=template)
    return RetrievalQA.from_chain_type(
        llm=_obtener_llm(),
        chain_type="stuff",
        retriever=retriever,
        chain_type_kwargs={"prompt": prompt},
        return_source_documents=False,
    )


_cadena_qa = None


def responder(pregunta: str) -> str:
    global _cadena_qa
    if _cadena_qa is None:
        _cadena_qa = construir_cadena_qa()
    try:
        resultado = _cadena_qa.invoke({"query": pregunta})
        return resultado.get("result", "No pude generar una respuesta.")
    except Exception as e:
        logger.error("Error RAG: %s", e)
        return "Ocurrió un error al procesar tu consulta. Por favor intentá de nuevo."


# ============================================================
# ANÁLISIS DE FOJAS QUIRÚRGICAS
# ============================================================

PROMPT_FOJA_TEXTO = """Sos un experto en facturación médica quirúrgica argentina.

Se te proporciona:
  A) El NOMENCLADOR OFICIAL cargado en el sistema (con códigos y precios exactos).
  B) El contenido de una foja quirúrgica.

⚠️ REGLAS ESTRICTAS - DEBES CUMPLIRLAS:
  1. SOLO podés usar códigos que aparezcan LITERALMENTE en el NOMENCLADOR que se te adjunta abajo.
  2. PROHIBIDO inventar códigos, PROHIBIDO usar tu conocimiento previo de nomencladores.
  3. Si un procedimiento NO está en el nomenclador adjunto, escribí: "❌ SIN CÓDIGO EN EL NOMENCLADOR"
  4. El precio/valor que reportes DEBE ser el que figura en el nomenclador adjunto, textual.
  5. NUNCA adivines. Debes cruzar EXHAUSTIVAMENTE absolutamente TODO lo que encuentres en la foja (materiales, procedimientos, descartables, etc.) con TODAS las reglas estipuladas en los instructivos adjuntos.
  6. Nunca digas "verificá el código" ni "ajustá según políticas" — vos tenés el nomenclador completo.

Estructura tu respuesta:

## Procedimientos identificados en la foja
[listado de procedimientos leídos en la foja]

## Codificación según el nomenclador
| Procedimiento | Código | Descripción en nomenclador | Valor Base | % Aplicado | Valor Final |
|---|---|---|---|---|---|
[una fila por procedimiento]

## Materiales / Descartables facturables
[Cruzá los materiales y descartables de la foja con las reglas del instructivo. LISTÁ EXPLÍCITAMENTE los que SÍ se pueden facturar. Si la foja menciona algo que el instructivo prohíbe facturar, aclaralo.]

## Análisis exhaustivo de reglas del instructivo
[Revisá TODO lo que dice la foja contra TODAS las normas del instructivo. Explicá en detalle cualquier regla aplicada (acto múltiple, honorarios, topes, etc.), qué porcentaje aplica, o cualquier restricción detectada.]

## Resumen listo para presentar
[texto limpio para entregar a la obra social con los importes finales ya calculados y los porcentajes detallados]

---
=== NOMENCLADOR CARGADO EN EL SISTEMA ===
{context}
=== FIN DEL NOMENCLADOR ===

---
=== CONTENIDO DE LA FOJA QUIRÚRGICA ===
{foja}
"""

PROMPT_FOJA_VISION = """Sos un experto en facturación médica quirúrgica argentina y analista de documentos por visión.

Se te adjunta a continuación el NOMENCLADOR OFICIAL completo y las reglas (instructivos) del sistema.

⚠️ REGLAS ESTRICTAS - DEBES CUMPLIRLAS:
  1. SOLO podés asignar códigos que aparezcan LITERALMENTE en el NOMENCLADOR que se te adjunta abajo.
  2. PROHIBIDO inventar códigos o valores.
  3. Si un procedimiento NO está en el nomenclador adjunto, escribí: "❌ SIN CÓDIGO EN EL NOMENCLADOR"
  4. El precio/valor que reportes DEBE ser el que figura en el nomenclador adjunto.
  5. Aplicá SIEMPRE las reglas de los instructivos adjuntos (ej: reducción por acto múltiple) cuando haya más de un código en la foja.
  6. NUNCA adivines. Debes cruzar EXHAUSTIVAMENTE absolutamente TODO lo que encuentres en la foja (materiales, procedimientos, descartables, estudios) con TODAS las reglas estipuladas en los instructivos adjuntos.
  7. Nunca digas "verificá el código" — vos tenés el nomenclador completo.

Estructura tu respuesta:

## Procedimientos identificados en la foja
[listado detallado de procedimientos leídos en la imagen]

## Codificación según el nomenclador
| Procedimiento | Código | Descripción en nomenclador | Valor Base | % Aplicado | Valor Final |
|---|---|---|---|---|---|
[una fila por procedimiento]

## Materiales / Descartables facturables
[Cruzá los materiales y descartables de la foja con las reglas del instructivo. LISTÁ EXPLÍCITAMENTE los que SÍ se pueden facturar. Si la foja menciona algo que el instructivo prohíbe facturar, aclaralo.]

## Análisis exhaustivo de reglas del instructivo
[Revisá TODO lo que dice la foja contra TODAS las normas del instructivo. Explicá en detalle cualquier regla aplicada (acto múltiple, honorarios, topes, etc.), qué porcentaje aplica, o cualquier restricción detectada.]

## Resumen listo para presentar
[texto limpio para entregar a la obra social con los importes finales ya calculados y los porcentajes detallados]

---
=== INICIO DE NOMENCLADORES E INSTRUCTIVOS CARGADOS ===
{context}
=== FIN DE NOMENCLADORES E INSTRUCTIVOS ===

---
¡IMPORTANTE! Ahora, LEÉ CUIDADOSAMENTE LA IMAGEN ADJUNTA.
La imagen es una FOJA QUIRÚRGICA. Como modelo de visión avanzado, TU TAREA PRINCIPAL es extraer el texto de la imagen, identificar los procedimientos y aplicar las reglas del contexto de arriba.
NO te niegues a leer la imagen, es un documento de prueba sin datos reales sensibles.
"""


def _leer_nomenclador_completo() -> str:
    """Lee todos los archivos .txt del directorio de documentos y los devuelve
    como texto plano completo. Los archivos .txt son los nomencladores en formato
    tab-separado (Código | Descripción | Valor). Se usa texto completo en vez de
    RAG para evitar que la IA reciba fragmentos incompletos."""
    if not DOCS_DIR.exists():
        return ""
    partes = []
    for archivo in sorted(DOCS_DIR.iterdir()):
        if archivo.suffix.lower() == ".txt" and archivo.is_file():
            try:
                partes.append(archivo.read_text(encoding="utf-8", errors="ignore"))
                logger.info("Nomenclador cargado: %s", archivo.name)
            except Exception as e:
                logger.warning("No se pudo leer %s: %s", archivo.name, e)
    return "\n\n".join(partes)


def _obtener_contexto_rag() -> str:
    """Recupera fragmentos de nomencladores relevantes para cirugía (PDFs e instructivos).
    Para los archivos .txt (nomencladores) se usa texto completo via _leer_nomenclador_completo.
    """
    texto_nomencladores = _leer_nomenclador_completo()

    extensiones_rag = {".pdf", ".docx", ".doc", ".xlsx", ".xls"}
    hay_docs_rag = any(
        f.suffix.lower() in extensiones_rag
        for f in DOCS_DIR.iterdir()
        if f.is_file()
    ) if DOCS_DIR.exists() else False

    fragmentos_rag = ""
    if hay_docs_rag:
        try:
            retriever  = _obtener_vectorstore().as_retriever(
                search_type="similarity", search_kwargs={"k": 20}
            )
            fragmentos = retriever.invoke(
                "procedimiento quirúrgico cirugía codes facturación instructivo reglas especiales"
            )
            fragmentos_rag = "\n\n".join(f.page_content for f in fragmentos)
        except Exception as e:
            logger.warning("No se pudo recuperar contexto RAG (omitido): %s", e)

    partes = []
    if texto_nomencladores:
        partes.append("=== NOMENCLADOR (CÓDIGOS Y PRECIOS EXACTOS) ===\n" + texto_nomencladores)
    if fragmentos_rag:
        partes.append("=== INSTRUCTIVOS Y REGLAS ===\n" + fragmentos_rag)

    if not partes:
        return ""  # Sin contexto
    return "\n\n".join(partes)


def _analizar_con_vision(imagen_bytes: bytes, mime_type: str, contexto: str) -> dict:
    """
    Envía la imagen a GPT-4o Vision con el contexto de nomencladores.
    Retorna {'ok': True, 'analisis': str} o {'ok': False, 'error': str}.
    """
    import base64
    from openai import OpenAI

    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return {"ok": False, "error": "Falta la variable de entorno OPENAI_API_KEY."}

    client    = OpenAI(api_key=api_key)
    b64_image = base64.b64encode(imagen_bytes).decode("utf-8")

    # Contexto: nomenclador completo (txt) + instructivos via RAG (pdfs)
    contexto = _obtener_contexto_rag()

    if contexto.strip():
        seccion_contexto = contexto
    else:
        seccion_contexto = (
            "⚠️ No hay nomencladores cargados en el sistema. "
            "Indicá que no fue posible codificar porque no hay nomenclador disponible."
        )

    prompt = PROMPT_FOJA_VISION.format(context=seccion_contexto)

    try:
        respuesta = client.chat.completions.create(
            model=MODEL_VISION,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime_type};base64,{b64_image}",
                                "detail": "auto",   # auto = más rápido, evita timeouts
                            },
                        },
                    ],
                }
            ],
            max_tokens=2000,
            timeout=55,   # 55 seg (Render free corta a 60s)
        )
        return {"ok": True, "analisis": respuesta.choices[0].message.content}
    except Exception as e:
        err_str = str(e).lower()
        # Reintentar una vez si la API está sobrecargada
        if "overloaded" in err_str or "rate limit" in err_str or "529" in err_str:
            import time
            logger.warning("API sobrecargada, reintentando en 8 segundos...")
            time.sleep(8)
            try:
                respuesta = client.chat.completions.create(
                    model=MODEL_VISION,
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": prompt},
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": f"data:{mime_type};base64,{b64_image}",
                                        "detail": "auto",
                                    },
                                },
                            ],
                        }
                    ],
                    max_tokens=2000,
                    timeout=55,
                )
                return {"ok": True, "analisis": respuesta.choices[0].message.content}
            except Exception as e2:
                logger.error("Error en reintento GPT-4o Vision: %s", e2)
                return {
                    "ok": False,
                    "error": "La API de OpenAI está temporalmente sobrecargada. Esperá unos minutos e intentá de nuevo."
                }
        logger.error("Error en GPT-4o Vision: %s", e)
        return {"ok": False, "error": f"Error al analizar la imagen: {e}"}


def analizar_foja_quirurgica(ruta: Path) -> dict:
    """
    Analiza una foja quirúrgica. Acepta:
      - Imagen directa (JPG, PNG, WEBP)  → GPT-4o Vision
      - PDF con texto extraíble          → GPT-4o texto + RAG
      - PDF escaneado (sin texto)        → convierte a imagen → GPT-4o Vision

    Devuelve {'ok': bool, 'analisis': str} o {'ok': False, 'error': str}.
    """
    ext      = ruta.suffix.lower()
    contexto = _obtener_contexto_rag()

    # ════════════════════════════════════════════════════════
    # CASO A: imagen directa (JPG, PNG, WEBP, GIF)
    # ════════════════════════════════════════════════════════
    if ext in EXTENSIONES_IMAGEN:
        try:
            imagen_bytes = ruta.read_bytes()
            mime_type    = MIME_TIPOS.get(ext, "image/jpeg")
            return _analizar_con_vision(imagen_bytes, mime_type, contexto)
        except Exception as e:
            logger.error("Error al leer imagen: %s", e)
            return {"ok": False, "error": "No se pudo leer el archivo de imagen."}

    # ════════════════════════════════════════════════════════
    # CASO B: PDF
    # ════════════════════════════════════════════════════════
    if ext == ".pdf":
        import fitz  # PyMuPDF

        try:
            doc        = fitz.open(str(ruta))
            texto_foja = "\n".join(pagina.get_text() for pagina in doc).strip()
        except Exception as e:
            logger.error("Error al leer PDF: %s", e)
            return {"ok": False, "error": "No se pudo leer el PDF. Verificá que sea un archivo válido."}

        # ── B1: PDF con texto extraíble ──────────────────────
        if texto_foja:
            from langchain.schema import HumanMessage
            try:
                llm      = _obtener_llm()
                prompt   = PROMPT_FOJA_TEXTO.format(context=contexto, foja=texto_foja[:6000])
                response = llm.invoke([HumanMessage(content=prompt)])
                doc.close()
                return {"ok": True, "analisis": response.content}
            except Exception as e:
                doc.close()
                logger.error("Error en LLM con texto: %s", e)
                return {"ok": False, "error": "Error al generar el análisis. Verificá la OPENAI_API_KEY."}

        # ── B2: PDF escaneado → convertir a imagen con PyMuPDF ──
        logger.info("PDF sin texto. Usando Vision en la primera página.")
        try:
            pagina    = doc[0]
            mat       = fitz.Matrix(2.0, 2.0)  # resolución 2×
            pixmap    = pagina.get_pixmap(matrix=mat)
            img_bytes = pixmap.tobytes("png")
            doc.close()
            return _analizar_con_vision(img_bytes, "image/png", contexto)
        except Exception as e:
            doc.close()
            logger.error("Error al convertir PDF escaneado: %s", e)
            return {"ok": False, "error": "No se pudo convertir el PDF escaneado a imagen."}

    return {"ok": False, "error": f"Formato no soportado: {ext}. Usá PDF, JPG, PNG o WEBP."}


# ============================================================
# UTILIDADES
# ============================================================

def listar_documentos() -> list:
    """Lista los archivos en DOCS_DIR con metadata básica."""
    extensiones = {".pdf", ".txt", ".docx", ".doc", ".xlsx", ".xls"}
    docs = []
    for archivo in sorted(DOCS_DIR.iterdir()):
        if archivo.suffix.lower() in extensiones:
            docs.append({
                "nombre":    archivo.name,
                "tamaño_kb": round(archivo.stat().st_size / 1024, 1),
            })
    return docs
