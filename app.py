import os
import time
import uuid
import threading
from pathlib import Path

import openpyxl
from flask import Flask, request, render_template, send_file, jsonify

# ── Módulos del bot de IA (opcionales: requieren pip install -r requirements.txt) ──
try:
    from bot_rag import (
        responder,
        agregar_documento_a_base,
        listar_documentos,
        analizar_foja_quirurgica,
        indexar_todos_los_documentos,
        DOCS_DIR,
    )
    from whatsapp_api import verificar_webhook, extraer_mensaje, enviar_texto
    BOT_DISPONIBLE = True
except ImportError as _e:
    BOT_DISPONIBLE = False
    _bot_error = (
        "El módulo de IA no está disponible. "
        "Ejecutá: pip install -r requirements.txt\n"
        f"Detalle: {_e}"
    )
    # Funciones vacías para que los endpoints no exploten
    def responder(p):              return _bot_error
    def agregar_documento_a_base(r): return 0
    def listar_documentos():       return []
    def analizar_foja_quirurgica(r): return {"ok": False, "error": _bot_error}
    def indexar_todos_los_documentos(): return 0
    def verificar_webhook(a):      return "Bot no disponible", 503
    def extraer_mensaje(p):        return None
    def enviar_texto(n, t):        return False
    DOCS_DIR = Path(__file__).resolve().parent / "documentos"
    DOCS_DIR.mkdir(exist_ok=True)

from sss_beneficiarios_hospitales.data import DataBeneficiariosSSSHospital


# ============================================================
# AUTO-INDEXING EN SEGUNDO PLANO
# ============================================================

def _iniciar_indexacion_background():
    if BOT_DISPONIBLE:
        def tarea():
            try:
                # Esto indexará los documentos si la colección está vacía
                indexar_todos_los_documentos()
            except Exception as e:
                print(f"Error en auto-indexación: {e}")
        
        t = threading.Thread(target=tarea, daemon=True)
        t.start()

# Disparar indexación al iniciar
_iniciar_indexacion_background()

# ============================================================
# CONFIGURACIÓN
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

UPLOAD_DIR = BASE_DIR / "uploads"
RESULT_DIR = BASE_DIR / "resultados"

UPLOAD_DIR.mkdir(exist_ok=True)
RESULT_DIR.mkdir(exist_ok=True)

FILA_INICIO = 11
COLUMNA_DNI = 1
COLUMNA_OBRA_SOCIAL = 7

# Antes: 0.5
# Ahora: 0.1 segundo entre consultas
ESPERA = 0.1

# Guardar Excel cada X consultas
GUARDAR_CADA = 50

app = Flask(__name__)

app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024


# ============================================================
# ESTADO DE PROCESAMIENTO
# ============================================================

procesos = {}


# ============================================================
# SSSALUD
# ============================================================

def crear_conexion_sss():

    usuario = os.environ.get("SSS_USER")
    password = os.environ.get("SSS_PASSWORD")

    if not usuario or not password:
        raise RuntimeError(
            "Faltan las variables SSS_USER y SSS_PASSWORD."
        )

    sss = DataBeneficiariosSSSHospital(
        user=usuario,
        password=password
    )

    # No guardar respuestas individuales
    sss._save_response = lambda filename, resp: None

    return sss


# ============================================================
# CONSULTA SSSALUD
# ============================================================

def consultar_dni(sss, dni):

    try:

        resultado = sss.query(dni)

        ok = resultado.get("ok")
        datos = resultado.get("resultados", {})

        if not ok:
            return "ERROR CONSULTA"

        afiliado = datos.get("afiliado")
        tablas = datos.get("tablas", [])

        # ----------------------------------------------------
        # NO AFILIADO
        # ----------------------------------------------------

        if afiliado is False:
            return "NO AFILIADO"

        # ----------------------------------------------------
        # AFILIADO
        # ----------------------------------------------------

        if afiliado is True:

            codigo_obra_social = None
            denominacion_obra_social = None

            for tabla in tablas:

                nombre_tabla = str(
                    tabla.get("name", "")
                ).strip().upper()

                if nombre_tabla != "AFILIADO":
                    continue

                data = tabla.get("data", {})

                codigo_obra_social = data.get(
                    "Código de Obra Social"
                )

                denominacion_obra_social = data.get(
                    "Denominación Obra Social"
                )

                break

            if codigo_obra_social and denominacion_obra_social:

                return (
                    f"{codigo_obra_social} - "
                    f"{denominacion_obra_social}"
                )

            if denominacion_obra_social:
                return str(denominacion_obra_social)

            if codigo_obra_social:
                return str(codigo_obra_social)

            return "AFILIADO - SIN OBRA SOCIAL IDENTIFICADA"

        return "ERROR CONSULTA"

    except Exception as e:

        print("ERROR CONSULTA:", e)

        return "ERROR CONSULTA"


# ============================================================
# PROCESAMIENTO EN SEGUNDO PLANO
# ============================================================

def procesar_archivo(
    id_proceso,
    archivo_entrada,
    archivo_salida
):

    estado = procesos[id_proceso]

    try:

        # ----------------------------------------------------
        # ABRIR EXCEL
        # ----------------------------------------------------

        wb = openpyxl.load_workbook(
            archivo_entrada
        )

        ws = wb.active

        ultima_fila = ws.max_row

        # ----------------------------------------------------
        # BUSCAR FILAS A PROCESAR
        #
        # IMPORTANTE:
        # AHORA VOLVEMOS A CONSULTAR TODAS LAS FILAS
        # QUE TENGAN DNI.
        #
        # NO SE SALTEAN POR TENER OBRA SOCIAL.
        # ----------------------------------------------------

        filas_a_procesar = []

        for fila in range(
            FILA_INICIO,
            ultima_fila + 1
        ):

            dni = ws.cell(
                fila,
                COLUMNA_DNI
            ).value

            if dni is None:
                continue

            dni = str(dni).strip()

            if dni == "":
                continue

            filas_a_procesar.append(fila)

        total = len(filas_a_procesar)

        estado["total"] = total
        estado["procesadas"] = 0
        estado["estado"] = "iniciando"

        print("")
        print("====================================================")
        print("       PROCESAMIENTO WEB SSSALUD")
        print("====================================================")
        print("Última fila:", ultima_fila)
        print("Total a consultar:", total)
        print("====================================================")

        # ----------------------------------------------------
        # CONEXIÓN
        # ----------------------------------------------------

        estado["estado"] = "conectando"

        sss = crear_conexion_sss()

        estado["estado"] = "consultando"

        # ----------------------------------------------------
        # CONTADORES
        # ----------------------------------------------------

        consultas = 0
        afiliados = 0
        no_afiliados = 0
        errores = 0

        inicio = time.time()

        # ----------------------------------------------------
        # PROCESAMIENTO
        # ----------------------------------------------------

        for indice, fila in enumerate(
            filas_a_procesar,
            start=1
        ):

            dni = ws.cell(
                fila,
                COLUMNA_DNI
            ).value

            dni = str(dni).strip()

            estado["fila"] = fila
            estado["dni"] = dni
            estado["estado"] = "consultando"

            print(
                f"[{indice}/{total}] "
                f"Fila {fila} | DNI {dni}"
            )

            # ------------------------------------------------
            # CONSULTA
            # ------------------------------------------------

            resultado_final = consultar_dni(
                sss,
                dni
            )

            # ------------------------------------------------
            # CONTADORES
            # ------------------------------------------------

            if resultado_final == "NO AFILIADO":

                no_afiliados += 1

            elif resultado_final == "ERROR CONSULTA":

                errores += 1

            else:

                afiliados += 1

            consultas += 1

            # ------------------------------------------------
            # ESCRIBIR RESULTADO
            # ------------------------------------------------

            ws.cell(
                fila,
                COLUMNA_OBRA_SOCIAL
            ).value = resultado_final

            # ------------------------------------------------
            # ACTUALIZAR ESTADO WEB
            # ------------------------------------------------

            estado["procesadas"] = indice
            estado["consultas"] = consultas
            estado["afiliados"] = afiliados
            estado["no_afiliados"] = no_afiliados
            estado["errores"] = errores
            estado["ultimo_resultado"] = resultado_final

            # ------------------------------------------------
            # PORCENTAJE
            # ------------------------------------------------

            if total > 0:

                estado["porcentaje"] = round(
                    (indice / total) * 100,
                    1
                )

            else:

                estado["porcentaje"] = 100

            # ------------------------------------------------
            # TIEMPO
            # ------------------------------------------------

            tiempo_transcurrido = time.time() - inicio

            estado["segundos"] = round(
                tiempo_transcurrido,
                1
            )

            if indice > 0:

                promedio = (
                    tiempo_transcurrido / indice
                )

                restantes = total - indice

                estado["estimado_restante"] = round(
                    promedio * restantes,
                    1
                )

            print(
                f"Resultado: {resultado_final}"
            )

            # ------------------------------------------------
            # GUARDADO PERIÓDICO
            # ------------------------------------------------

            if consultas % GUARDAR_CADA == 0:

                wb.save(
                    archivo_salida
                )

                print(
                    ">>> PROGRESO GUARDADO"
                )

            # ------------------------------------------------
            # ESPERA PEQUEÑA
            # ------------------------------------------------

            if ESPERA > 0:

                time.sleep(
                    ESPERA
                )

        # ----------------------------------------------------
        # GUARDADO FINAL
        # ----------------------------------------------------

        estado["estado"] = "guardando"

        wb.save(
            archivo_salida
        )

        # ----------------------------------------------------
        # FINALIZADO
        # ----------------------------------------------------

        estado["estado"] = "terminado"

        estado["procesadas"] = total
        estado["porcentaje"] = 100
        estado["archivo"] = Path(
            archivo_salida
        ).name

        estado["segundos"] = round(
            time.time() - inicio,
            1
        )

        estado["estimado_restante"] = 0

        print("")
        print("====================================================")
        print("              PROCESO TERMINADO")
        print("====================================================")
        print("Consultas:", consultas)
        print("Afiliados:", afiliados)
        print("No afiliados:", no_afiliados)
        print("Errores:", errores)
        print("====================================================")

    except Exception as e:

        print("")
        print("====================================================")
        print("ERROR GENERAL")
        print("====================================================")
        print(e)

        estado["estado"] = "error"
        estado["error"] = str(e)


# ============================================================
# PÁGINA PRINCIPAL
# ============================================================

@app.route("/", methods=["GET"])
def index():

    return render_template(
        "index.html"
    )


# ============================================================
# INICIAR PROCESAMIENTO
# ============================================================

@app.route(
    "/procesar",
    methods=["POST"]
)
def procesar():

    archivo = request.files.get(
        "archivo"
    )

    # --------------------------------------------------------
    # VALIDAR ARCHIVO
    # --------------------------------------------------------

    if not archivo or archivo.filename == "":

        return """
        <h2>No se seleccionó ningún archivo.</h2>
        <a href="/">Volver</a>
        """

    if not archivo.filename.lower().endswith(
        ".xlsx"
    ):

        return """
        <h2>El archivo debe ser .xlsx</h2>
        <a href="/">Volver</a>
        """

    # --------------------------------------------------------
    # ID ÚNICO
    # --------------------------------------------------------

    identificador = uuid.uuid4().hex

    archivo_entrada = (
        UPLOAD_DIR
        /
        f"{identificador}_{archivo.filename}"
    )

    archivo_salida = (
        RESULT_DIR
        /
        f"procesado_{identificador}_{archivo.filename}"
    )

    # --------------------------------------------------------
    # GUARDAR ARCHIVO
    # --------------------------------------------------------

    archivo.save(
        archivo_entrada
    )

    # --------------------------------------------------------
    # CREAR ESTADO
    # --------------------------------------------------------

    procesos[identificador] = {

        "estado": "preparando",

        "total": 0,

        "procesadas": 0,

        "fila": 0,

        "dni": "",

        "consultas": 0,

        "afiliados": 0,

        "no_afiliados": 0,

        "errores": 0,

        "ultimo_resultado": "",

        "porcentaje": 0,

        "segundos": 0,

        "estimado_restante": 0,

        "archivo": "",

        "error": ""
    }

    # --------------------------------------------------------
    # INICIAR HILO
    # --------------------------------------------------------

    hilo = threading.Thread(

        target=procesar_archivo,

        args=(
            identificador,
            archivo_entrada,
            archivo_salida
        ),

        daemon=True
    )

    hilo.start()

    # --------------------------------------------------------
    # MOSTRAR PANTALLA DE PROGRESO
    # --------------------------------------------------------

    return render_template(
        "progreso.html",
        id_proceso=identificador
    )


# ============================================================
# CONSULTAR PROGRESO
# ============================================================

@app.route(
    "/progreso/<id_proceso>"
)
def progreso(id_proceso):

    estado = procesos.get(
        id_proceso
    )

    if not estado:

        return jsonify({

            "estado": "error",

            "error": "Proceso no encontrado"

        }), 404

    return jsonify(
        estado
    )


# ============================================================
# DESCARGAR RESULTADO
# ============================================================

@app.route(
    "/descargar/<nombre>"
)
def descargar(nombre):

    archivo = (
        RESULT_DIR
        /
        nombre
    )

    if not archivo.exists():

        return (
            "Archivo no encontrado",
            404
        )

    return send_file(
        archivo,
        as_attachment=True
    )


# ============================================================
# WEBHOOK WHATSAPP — VERIFICACIÓN (GET)
# ============================================================

@app.route("/webhook", methods=["GET"])
def webhook_verificar():

    challenge, status = verificar_webhook(request.args)
    return challenge, status


# ============================================================
# WEBHOOK WHATSAPP — MENSAJES ENTRANTES (POST)
# ============================================================

@app.route("/webhook", methods=["POST"])
def webhook_recibir():

    payload = request.get_json(silent=True) or {}

    mensaje = extraer_mensaje(payload)

    if mensaje:

        numero = mensaje["numero"]
        texto  = mensaje["texto"]

        # Procesamos en hilo separado para no bloquear la respuesta a Meta
        def responder_async():
            respuesta = responder(texto)
            enviar_texto(numero, respuesta)

        hilo = threading.Thread(target=responder_async, daemon=True)
        hilo.start()

    # Meta requiere siempre un 200 rápido
    return jsonify({"status": "ok"}), 200


# ============================================================
# Estado de indexado en memoria
_estado_indexado = {}


@app.route("/api/upload_docs", methods=["POST"])
def upload_documento():

    archivo = request.files.get("documento")

    if not archivo or archivo.filename == "":
        return jsonify({"ok": False, "error": "No se recibió ningún archivo."}), 400

    extensiones_permitidas = {".pdf", ".txt", ".docx", ".doc", ".xlsx", ".xls"}
    ext = Path(archivo.filename).suffix.lower()

    if ext not in extensiones_permitidas:
        return jsonify({
            "ok": False,
            "error": "Formato no permitido. Usá: PDF, TXT, DOCX o XLSX."
        }), 400

    destino = DOCS_DIR / archivo.filename
    archivo.save(destino)

    nombre = archivo.filename
    _estado_indexado[nombre] = {"estado": "procesando", "fragmentos": 0}

    # Indexar en segundo plano para no bloquear la respuesta HTTP
    def _indexar():
        try:
            n = agregar_documento_a_base(destino)
            _estado_indexado[nombre] = {"estado": "listo", "fragmentos": n}
        except Exception as e:
            _estado_indexado[nombre] = {"estado": "error", "error": str(e)}

    threading.Thread(target=_indexar, daemon=True).start()

    # Respuesta inmediata — el cliente puede consultar /api/estado_indexado/<nombre>
    return jsonify({
        "ok":      True,
        "nombre":  nombre,
        "mensaje": "Archivo recibido. Indexando en segundo plano...",
    })


@app.route("/api/estado_indexado/<nombre>")
def estado_indexado(nombre):
    estado = _estado_indexado.get(nombre, {"estado": "desconocido"})
    return jsonify(estado)


# ============================================================
# LISTAR DOCUMENTOS DEL BOT
# ============================================================

@app.route("/api/docs", methods=["GET"])
def listar_docs():
    return jsonify(listar_documentos())


# ============================================================
# ANÁLISIS DE FOJA QUIRÚRGICA
# ============================================================

FOJAS_DIR = BASE_DIR / "fojas_tmp"
FOJAS_DIR.mkdir(exist_ok=True)


@app.route("/api/analizar_foja", methods=["POST"])
def analizar_foja():

    archivo = request.files.get("foja")

    if not archivo or archivo.filename == "":
        return jsonify({"ok": False, "error": "No se recibió ningún archivo."}), 400

    EXTENSIONES_OK = {".pdf", ".jpg", ".jpeg", ".png", ".webp", ".gif"}
    ext = Path(archivo.filename).suffix.lower()

    if ext not in EXTENSIONES_OK:
        return jsonify({
            "ok": False,
            "error": "Formato no soportado. Usá PDF, JPG, PNG o WEBP."
        }), 400

    # Guardar temporalmente con extensión original para que bot_rag la detecte
    nombre_tmp = "{}_{}".format(uuid.uuid4().hex, archivo.filename)
    ruta_tmp   = FOJAS_DIR / nombre_tmp
    archivo.save(ruta_tmp)

    try:
        resultado = analizar_foja_quirurgica(ruta_tmp)
    finally:
        # Eliminar el temporal después del análisis
        try:
            ruta_tmp.unlink(missing_ok=True)
        except Exception:
            pass

    return jsonify(resultado)


# ============================================================
# EJECUTAR
# ============================================================

if __name__ == "__main__":

    app.run(

        host="127.0.0.1",

        port=5000,

        debug=False
    )